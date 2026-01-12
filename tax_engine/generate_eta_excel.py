#!/usr/bin/env python3
"""
Generate ETA-compliant Excel files for VAT returns
Uses official ETA templates for sales and purchases
"""

import openpyxl
from openpyxl import load_workbook
import json
import os
from datetime import datetime
from shutil import copy2

# Load templates
SALES_TEMPLATE = 'VAT Low/sales-doc_0.xlsx'
PURCHASES_TEMPLATE = 'VAT Low/مستندات المشتريات_0.xlsx'
OUTPUT_DIR = 'output/excel_uploads_eta'

# Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Get all VAT return JSON files
vat_files = sorted([f for f in os.listdir('output') if f.endswith('_vat_return.json')])

print('📊 جاري إنشاء ملفات Excel بصيغة المصلحة...\n')

for vat_file in vat_files:
    # Load VAT data
    with open(f'output/{vat_file}', 'r', encoding='utf-8') as f:
        vat_data = json.load(f)
    
    period = vat_data['period']
    period_name = vat_data['periodName']
    
    print(f'⚙️  {period_name}...')
    
    # ===== SALES FILE =====
    wb_sales = load_workbook(SALES_TEMPLATE)
    ws_sales = wb_sales.active
    
    row_num = 3  # Start after header rows
    for invoice in vat_data['sales']['local']['items']:
        # Calculate net amount (before tax)
        net_amount = invoice['total'] - invoice['vat']
        
        ws_sales.cell(row_num, 1).value = 'I'  # Invoice type
        ws_sales.cell(row_num, 2).value = 'T1'  # VAT type
        ws_sales.cell(row_num, 3).value = ''  # Schedule goods type (empty for general)
        ws_sales.cell(row_num, 4).value = invoice['id']  # Invoice number
        ws_sales.cell(row_num, 5).value = invoice['customer']  # Customer name
        ws_sales.cell(row_num, 6).value = ''  # Customer tax reg (unknown)
        ws_sales.cell(row_num, 7).value = ''  # Customer file number
        ws_sales.cell(row_num, 8).value = 'Egypt'  # Address
        ws_sales.cell(row_num, 9).value = ''  # National ID
        ws_sales.cell(row_num, 10).value = ''  # Phone
        ws_sales.cell(row_num, 11).value = invoice['date']  # Invoice date
        ws_sales.cell(row_num, 12).value = 'Electronics'  # Product name
        ws_sales.cell(row_num, 13).value = ''  # Product code
        ws_sales.cell(row_num, 14).value = 'GS1'  # Statement type
        ws_sales.cell(row_num, 15).value = 'GS1'  # Commodity type
        ws_sales.cell(row_num, 16).value = 'EA'  # Unit of measure
        ws_sales.cell(row_num, 17).value = net_amount  # Unit price (net)
        ws_sales.cell(row_num, 18).value = 'T1'  # Tax category (14%)
        ws_sales.cell(row_num, 19).value = 1  # Quantity
        ws_sales.cell(row_num, 20).value = net_amount  # Total amount
        ws_sales.cell(row_num, 21).value = 0  # Discount
        ws_sales.cell(row_num, 22).value = net_amount  # Net amount
        ws_sales.cell(row_num, 23).value = invoice['vat']  # Tax value
        ws_sales.cell(row_num, 24).value = invoice['total']  # Total
        
        row_num += 1
    
    # Save sales file
    sales_filename = f'{OUTPUT_DIR}/{period}_Sales.xlsx'
    wb_sales.save(sales_filename)
    
    # ===== PURCHASES FILE =====
    wb_purch = load_workbook(PURCHASES_TEMPLATE)
    ws_purch = wb_purch.active
    
    row_num = 3  # Start after header rows
    for invoice in vat_data['inputs']['items']:
        # Calculate net amount (before tax)
        net_amount = invoice['total'] - invoice['vat']
        
        ws_purch.cell(row_num, 1).value = 'I'  # Invoice type
        ws_purch.cell(row_num, 2).value = 'T1'  # VAT type
        ws_purch.cell(row_num, 3).value = ''  # Schedule goods type
        ws_purch.cell(row_num, 4).value = invoice['id']  # Invoice number
        ws_purch.cell(row_num, 5).value = invoice['customer']  # Supplier name
        ws_purch.cell(row_num, 6).value = ''  # Supplier tax reg
        ws_purch.cell(row_num, 7).value = ''  # Supplier file number
        ws_purch.cell(row_num, 8).value = 'Egypt'  # Address
        ws_purch.cell(row_num, 9).value = ''  # National ID
        ws_purch.cell(row_num, 10).value = ''  # Phone
        ws_purch.cell(row_num, 11).value = invoice['date']  # Invoice date
        ws_purch.cell(row_num, 12).value = 'Electronics'  # Product name
        ws_purch.cell(row_num, 13).value = ''  # Product code
        ws_purch.cell(row_num, 14).value = 'GS1'  # Statement type
        ws_purch.cell(row_num, 15).value = 'GS1'  # Commodity type
        ws_purch.cell(row_num, 16).value = 'EA'  # Unit of measure
        ws_purch.cell(row_num, 17).value = net_amount  # Unit price
        ws_purch.cell(row_num, 18).value = 'T1'  # Tax category (14%)
        ws_purch.cell(row_num, 19).value = 1  # Quantity
        ws_purch.cell(row_num, 20).value = net_amount  # Total amount
        ws_purch.cell(row_num, 21).value = 0  # Discount
        ws_purch.cell(row_num, 22).value = net_amount  # Net amount
        ws_purch.cell(row_num, 23).value = invoice['vat']  # Tax value
        ws_purch.cell(row_num, 24).value = invoice['total']  # Total
        
        row_num += 1
    
    # Save purchases file
    purch_filename = f'{OUTPUT_DIR}/{period}_Purchases.xlsx'
    wb_purch.save(purch_filename)
    
    # Status
    status_icon = '💚' if vat_data['summary']['status'] == 'Refundable' else '💙'
    status_text = 'رصيد دائن' if vat_data['summary']['status'] == 'Refundable' else 'مستحق'
    print(f'   {status_icon} Sales: {len(vat_data["sales"]["local"]["items"])} فواتير')
    print(f'   {status_icon} Purchases: {len(vat_data["inputs"]["items"])} فواتير')
    print(f'   {status_icon} Net: {abs(vat_data["summary"]["netVATDue"]):.2f} ج.م ({status_text})\n')

print(f'✅ تم إنشاء {len(vat_files) * 2} ملف Excel بنجاح!')
print(f'📁 الملفات موجودة في: {OUTPUT_DIR}')
print('\n💡 كل شهر له ملفين:')
print('   - YYYY-MM_Sales.xlsx (فواتير المبيعات)')
print('   - YYYY-MM_Purchases.xlsx (فواتير المشتريات)')
